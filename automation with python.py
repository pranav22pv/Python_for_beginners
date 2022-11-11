import openpyxl as xl
from openpyxl.chart import BarChart, Reference    # we have a module called chart in openpyxl , from that module we
#  are importing 2 classes - BarChart and Reference

def process_workbook(filename):     # bfr converting this below code as a function instead of filename transaction.xlsx was used
    wb = xl.load_workbook(filename)    #  reads the excel sheet contents and stores in wb
    sheet = wb['Sheet1']   # opening  sheet 1 and storing in sheet
    # accessing a particular cell in the spreadsheet
    # 2 methods
    cell = sheet['a1']  # coordinate is a1 which is nothing but the combination of column and row
    # or
    cell = sheet.cell(1,1)  #  (row,column)

    print(cell.value)


    # to know how many rows we have in the spreadsheet:
    #sheet.max_row
    print(sheet.max_row)
    print('\n')


    for row in range(2,sheet.max_row+1):    # starting range from 2 because 1st row is attribute names and value starts from 2nd row
        cell = sheet.cell(row,3)    # column we wanna change is 3. so (row,column) = (row, 3)
        print(cell.value)
        corrected_price = cell.value*0.9
        corrected_price_cell = sheet.cell(row,4)
        corrected_price_cell.value = corrected_price


    values = Reference(sheet,min_row=2,max_row = sheet.max_row,min_col = 4,max_col = 4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')
    wb.save(filename)
    # wb.save('transactions2.xlsx')
    # wb.save('transactions3.xlsx')    # thi has been put in comment bcoz we can only run this code once, 2nd time shows error